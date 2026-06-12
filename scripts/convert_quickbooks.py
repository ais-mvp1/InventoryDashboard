"""
Convert raw QuickBooks report exports into a Parts_Tracker workbook for the dashboard.

Input: a month folder (e.g. quickbook/2026-04) containing three QuickBooks exports:
  - Purchases by Product/Service Detail  (filename contains "Purchases")
  - Sales by Product/Service Detail      (filename contains "Sales by Product")
  - Sales by Class Detail                (filename contains "Sales by Class")

Output: Parts_Tracker_<Mon><Year>.xlsx in the same folder, with the same three sheets
the dashboard upload expects: Summary, Parts Detail, Install Tracker.

Usage:
  python scripts/convert_quickbooks.py quickbook/2026-04
  python scripts/convert_quickbooks.py --all          # convert every month folder
"""
from __future__ import annotations

import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
QB_DIR = ROOT / "quickbook"

DASH = "\u2014"  # em dash placeholder used for empty cells
COMPANY = "NTL Express"

# Sold line items that are services/fees, not trackable parts.
EXCLUDED_ITEMS = {
    "labor",
    "labor - diagnosic",
    "parts",
    "delivery",
    "flat tire",
    "tire disposal",
    "supply",
    "credit card -fee",
    "surcharge -fee",
    "discount.",
    "sales tax",
    "fet",
    "ca tire fee",
    "ca battery fee",
    "regulatory compliance fee",
    "state oil recycling tax",
    "not specified",
}

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


def is_excluded(name: str) -> bool:
    return re.sub(r"\s+", " ", name).strip().lower() in EXCLUDED_ITEMS


def norm_code(name: str) -> str:
    """Group names read from numeric cells come back as '133510.0' — strip the .0."""
    s = str(name).strip()
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s


def parse_date(v) -> datetime | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fmt_date(d: datetime | None) -> str | None:
    return d.strftime("%m/%d/%Y") if d else None


def to_num(v) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def clean_int_str(v) -> str | None:
    """Invoice/bill numbers: render 3419.0 -> '3419'."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip() or None


def parse_grouped(path: Path, header_row: int = 4) -> list[tuple[tuple[str, ...], dict]]:
    """Parse a QuickBooks grouped detail report.

    Group names live in column A (possibly nested, closed by 'Total for X' rows);
    detail rows have column A empty. Returns (group_path, row_dict) pairs.
    """
    raw = pd.read_excel(path, header=None)
    headers = [str(h) if pd.notna(h) else f"col{i}" for i, h in enumerate(raw.iloc[header_row])]
    out: list[tuple[tuple[str, ...], dict]] = []
    stack: list[str] = []
    for r in range(header_row + 1, len(raw)):
        row = raw.iloc[r]
        c0 = row.iloc[0]
        if pd.notna(c0):
            s = str(c0).strip()
            low = s.lower()
            if low.startswith("total for "):
                name = s[len("total for "):].strip()
                while stack and stack[-1] != name:
                    stack.pop()
                if stack:
                    stack.pop()
            elif low == "total":
                stack = []
            elif re.search(r"accrual basis|cash basis", low):
                pass  # report footer
            else:
                stack.append(s)
            continue
        if stack and row.iloc[1:].notna().any():
            out.append((tuple(stack), {headers[i]: row.iloc[i] for i in range(1, len(headers))}))
    return out


@dataclass
class PartAgg:
    code: str
    description: str | None = None
    vendors: list[str] = field(default_factory=list)
    bills: list[str] = field(default_factory=list)
    purchase_dates: list[datetime] = field(default_factory=list)
    qty_purchased: float = 0.0
    total_cost: float = 0.0
    rate_qty: float = 0.0  # sum of rate*qty over bill rows, for the unit-cost average
    rate_qty_total: float = 0.0
    qty_sold: float = 0.0
    revenue: float = 0.0
    sale_dates: list[datetime] = field(default_factory=list)
    invoices: list[str] = field(default_factory=list)
    trucks: list[str] = field(default_factory=list)

    @property
    def unit_cost(self) -> float:
        # Qty-weighted average bill rate; credits with negative amounts (e.g. core
        # returns) keep their positive per-unit rate this way.
        if self.rate_qty_total:
            return round(self.rate_qty / self.rate_qty_total, 2)
        return round(self.total_cost / self.qty_purchased, 2) if self.qty_purchased else 0.0

    @property
    def status(self) -> str:
        if self.qty_sold == 0:
            return "In Stock"
        return "Fully Installed" if self.qty_sold >= self.qty_purchased else "Partial"


@dataclass
class SaleLine:
    part: str
    description: str | None
    date: datetime | None
    invoice: str | None
    qty: float
    amount: float
    truck: str


def add_unique(lst: list, item) -> None:
    if item and item not in lst:
        lst.append(item)


def find_report(folder: Path, *keywords: str) -> Path:
    for f in sorted(folder.glob("*.xlsx")):
        name = f.name.replace("+", " ").lower()
        if all(k.lower() in name for k in keywords):
            return f
    raise FileNotFoundError(f"No report matching {keywords} in {folder}")


def convert_month(folder: Path) -> Path:
    purch_f = find_report(folder, "purchases")
    sales_item_f = find_report(folder, "sales by product")
    sales_class_f = find_report(folder, "sales by class")

    parts: dict[str, PartAgg] = {}

    def get_part(code: str) -> PartAgg:
        return parts.setdefault(code, PartAgg(code=code))

    # ---- Purchases: only real Bill / Vendor Credit lines with an item group
    for path, row in parse_grouped(purch_f):
        leaf = norm_code(path[-1])
        ttype = str(row.get("Transaction Type") or "")
        if is_excluded(leaf) or leaf.lower() == "not specified":
            continue
        if ttype not in ("Bill", "Vendor Credit"):
            continue
        qty = to_num(row.get("Qty"))
        amount = to_num(row.get("Amount"))
        if qty == 0 and amount == 0:
            continue
        p = get_part(leaf)
        if not p.description:
            p.description = clean_int_str(row.get("Memo/Description"))
        add_unique(p.vendors, clean_int_str(row.get("Vendor")))
        add_unique(p.bills, clean_int_str(row.get("Num")))
        d = parse_date(row.get("Date"))
        if d:
            p.purchase_dates.append(d)
        # Quantities/amounts in the report are already signed (vendor credits negative).
        p.qty_purchased += qty
        p.total_cost += amount
        if ttype == "Bill" and qty > 0:
            p.rate_qty += to_num(row.get("Rate")) * qty
            p.rate_qty_total += qty

    # ---- Sales: class report gives the truck; product report catches classless sales
    sale_lines: list[SaleLine] = []
    class_keys: Counter = Counter()

    for path, row in parse_grouped(sales_class_f):
        item_full = row.get("Product/Service full name")
        if pd.isna(item_full):
            continue
        leaf = norm_code(str(item_full).split(":")[-1])
        if is_excluded(leaf):
            continue
        truck = path[-1]
        truck = "Others" if truck.lower() == "not specified" else truck
        date = parse_date(row.get("Transaction date"))
        inv = clean_int_str(row.get("Num"))
        amount = round(to_num(row.get("Amount")), 2)
        sale_lines.append(SaleLine(
            part=leaf,
            description=clean_int_str(row.get("Description")),
            date=date,
            invoice=inv,
            qty=to_num(row.get("Quantity")),
            amount=amount,
            truck=truck,
        ))
        class_keys[(leaf, inv, fmt_date(date), amount)] += 1

    for path, row in parse_grouped(sales_item_f):
        leaf = norm_code(path[-1])
        if is_excluded(leaf):
            continue
        date = parse_date(row.get("Transaction date"))
        inv = clean_int_str(row.get("Num"))
        amount = round(to_num(row.get("Amount")), 2)
        key = (leaf, inv, fmt_date(date), amount)
        if class_keys[key] > 0:
            class_keys[key] -= 1  # already captured with a truck from the class report
            continue
        sale_lines.append(SaleLine(
            part=leaf,
            description=clean_int_str(row.get("Description")),
            date=date,
            invoice=inv,
            qty=to_num(row.get("Quantity")),
            amount=amount,
            truck="Others",
        ))

    for line in sale_lines:
        p = get_part(line.part)
        if not p.description:
            p.description = line.description
        p.qty_sold += line.qty
        p.revenue += line.amount
        if line.date:
            add_unique(p.sale_dates, line.date)
        add_unique(p.invoices, line.invoice)
        add_unique(p.trucks, line.truck)

    if not parts:
        raise RuntimeError(f"No parts found in {folder} — check the report files.")

    # ---- Period label from actual transaction dates
    all_dates = [d for p in parts.values() for d in p.purchase_dates + p.sale_dates]
    start, end = min(all_dates), max(all_dates)
    period = (
        f"{MONTHS[start.month - 1]} {start.day} {DASH} "
        f"{MONTHS[end.month - 1]} {end.day}, {end.year}"
    )

    # ---- Output filename from the dominant month
    month_counts = Counter((d.year, d.month) for d in all_dates)
    (year, month), _ = month_counts.most_common(1)[0]
    out_path = folder / f"Parts_Tracker_{MONTHS[month - 1][:3]}{year}.xlsx"

    write_workbook(out_path, parts, sale_lines, period)
    return out_path


def write_workbook(out_path: Path, parts: dict[str, PartAgg],
                   sale_lines: list[SaleLine], period: str) -> None:
    wb = Workbook()
    bold = Font(bold=True)
    ordered = sorted(parts.values(), key=lambda p: p.code.lower())

    # ---- Summary
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = f"{COMPANY} {DASH} Parts Purchase & Install Tracker"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B3"] = period
    kpi_headers = ["Total Parts", "Fully Installed", "In Stock", "Total Cost", "Total Revenue"]
    for i, h in enumerate(kpi_headers):
        c = ws.cell(row=5, column=2 + i, value=h)
        c.font = bold
    statuses = ["Fully Installed", "Partial", "In Stock"]
    by_status = {s: [p for p in ordered if p.status == s] for s in statuses}
    ws.cell(row=6, column=2, value=len(ordered))
    ws.cell(row=6, column=3, value=len(by_status["Fully Installed"]))
    ws.cell(row=6, column=4, value=len(by_status["In Stock"]))
    ws.cell(row=6, column=5, value=round(sum(p.total_cost for p in ordered), 2))
    ws.cell(row=6, column=6, value=round(sum(p.revenue for p in ordered), 2))
    table_headers = ["Status", "# Parts", "Qty Purchased", "Qty Sold", "Total Cost", "Revenue"]
    for i, h in enumerate(table_headers):
        c = ws.cell(row=9, column=2 + i, value=h)
        c.font = bold
    for r, s in enumerate(statuses, start=10):
        grp = by_status[s]
        ws.cell(row=r, column=2, value=s)
        ws.cell(row=r, column=3, value=len(grp))
        ws.cell(row=r, column=4, value=round(sum(p.qty_purchased for p in grp), 2))
        ws.cell(row=r, column=5, value=round(sum(p.qty_sold for p in grp), 2))
        ws.cell(row=r, column=6, value=round(sum(p.total_cost for p in grp), 2))
        ws.cell(row=r, column=7, value=round(sum(p.revenue for p in grp), 2))
    ws.cell(row=13, column=2, value="TOTAL").font = bold
    for col in "CDEFG":
        ws[f"{col}13"] = f"=SUM({col}10:{col}12)"

    # ---- Parts Detail
    ws = wb.create_sheet("Parts Detail")
    ws["A1"] = f"Parts {DASH} Purchased & Sold ({period})"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Part Code", "Description", "Vendor", "Bill #", "Purchase Date",
               "Qty Purchased", "Unit Cost", "Total Cost", "Qty Sold", "Sale Date(s)",
               "Invoice #(s)", "Truck / Trailer", "Revenue", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h).font = bold

    def num(v: float):
        v = round(v, 2)
        return int(v) if v == int(v) else v

    for r, p in enumerate(ordered, start=3):
        has_purchase = bool(p.purchase_dates or p.qty_purchased)
        sale_dates = ", ".join(fmt_date(d) for d in sorted(p.sale_dates))
        ws.cell(row=r, column=1, value=p.code)
        ws.cell(row=r, column=2, value=p.description or DASH)
        ws.cell(row=r, column=3, value=", ".join(p.vendors) if p.vendors else DASH)
        ws.cell(row=r, column=4, value=", ".join(p.bills) if p.bills else DASH)
        ws.cell(row=r, column=5, value=fmt_date(min(p.purchase_dates)) if p.purchase_dates else DASH)
        ws.cell(row=r, column=6, value=num(p.qty_purchased) if has_purchase else 0)
        ws.cell(row=r, column=7, value=num(p.unit_cost))
        ws.cell(row=r, column=8, value=num(p.total_cost))
        ws.cell(row=r, column=9, value=num(p.qty_sold))
        ws.cell(row=r, column=10, value=sale_dates or DASH)
        ws.cell(row=r, column=11, value=", ".join(sorted(p.invoices, key=lambda x: (len(x), x))) or DASH)
        ws.cell(row=r, column=12, value=", ".join(sorted(p.trucks)) or DASH)
        ws.cell(row=r, column=13, value=num(p.revenue))
        ws.cell(row=r, column=14, value=p.status)

    widths = [24, 32, 28, 16, 13, 13, 10, 10, 9, 24, 16, 20, 10, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    # ---- Install Tracker
    ws = wb.create_sheet("Install Tracker")
    ws["A1"] = f"Install Tracker {DASH} Parts Sold & Installed ({period})"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (f"Purchase date can be used for warranty reference {DASH} "
                "use Bill # to confirm with vendor.")
    it_headers = ["Truck / Trailer", "Part Code", "Description", "Vendor", "Bill #",
                  "Purchase Date", "Install Date", "Invoice #", "Unit Cost"]
    for i, h in enumerate(it_headers, start=1):
        ws.cell(row=3, column=i, value=h).font = bold

    by_truck: dict[str, list[SaleLine]] = {}
    for line in sale_lines:
        by_truck.setdefault(line.truck, []).append(line)

    r = 4
    for truck in sorted(by_truck, key=str.lower):
        ws.cell(row=r, column=1, value=f"  {truck}").font = bold
        r += 1
        lines = sorted(by_truck[truck], key=lambda l: (l.date or datetime.min, l.part.lower()))
        for line in lines:
            p = parts[line.part]
            ws.cell(row=r, column=1, value=truck)
            ws.cell(row=r, column=2, value=line.part)
            ws.cell(row=r, column=3, value=line.description or p.description or DASH)
            ws.cell(row=r, column=4, value=", ".join(p.vendors) if p.vendors else DASH)
            ws.cell(row=r, column=5, value=", ".join(p.bills) if p.bills else DASH)
            ws.cell(row=r, column=6,
                    value=fmt_date(min(p.purchase_dates)) if p.purchase_dates else DASH)
            ws.cell(row=r, column=7, value=fmt_date(line.date) or DASH)
            ws.cell(row=r, column=8, value=line.invoice or DASH)
            ws.cell(row=r, column=9, value=p.unit_cost)
            r += 1

    widths = [18, 24, 32, 28, 16, 13, 12, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w

    wb.save(out_path)


def main() -> None:
    args = sys.argv[1:]
    if args and args[0] == "--all":
        folders = sorted(d for d in QB_DIR.iterdir() if d.is_dir() and re.fullmatch(r"\d{4}-\d{2}", d.name))
    elif args:
        folders = [Path(args[0])]
    else:
        print("Usage: python scripts/convert_quickbooks.py <month-folder | --all>", file=sys.stderr)
        sys.exit(1)

    for folder in folders:
        if not folder.is_dir():
            print(f"Not a folder: {folder}", file=sys.stderr)
            sys.exit(1)
        out = convert_month(folder)
        print(f"Wrote {out}")


if __name__ == "__main__":
    main()
